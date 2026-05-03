import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from scipy import stats
from scipy.optimize import curve_fit
from datetime import datetime, timedelta, date
import openpyxl
import io, os, zipfile, re
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(page_title="GE Vernova · Desgaste", page_icon="⚡",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
  [data-testid="stAppViewContainer"]{background:#0a0e1a;color:#e8eaf0}
  [data-testid="stSidebar"]{background:#0d1220;border-right:1px solid #1e2a45}
  [data-testid="stSidebar"] *{color:#c8d0e0!important}
  .ge-header{display:flex;align-items:center;gap:18px;padding:18px 28px;margin-bottom:6px;
    background:linear-gradient(90deg,#0d1220 0%,#0f2040 60%,#0d1220 100%);
    border-bottom:2px solid #1565c0;border-radius:0 0 10px 10px}
  .ge-header h1{margin:0;font-size:1.55rem;font-weight:700;color:#e8f0fe;letter-spacing:.03em}
  .ge-header span{font-size:.85rem;color:#7986cb;font-style:italic}
  .kpi-grid{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:20px}
  .kpi-card{flex:1;min-width:140px;background:#111827;border:1px solid #1e3a5f;
    border-radius:10px;padding:14px 18px;box-shadow:0 2px 12px rgba(0,0,0,.4)}
  .kpi-label{font-size:.72rem;color:#7986cb;text-transform:uppercase;letter-spacing:.08em;margin-bottom:4px}
  .kpi-value{font-size:1.5rem;font-weight:700;color:#e8f0fe}
  .kpi-sub{font-size:.75rem;color:#546e8a;margin-top:2px}
  .kpi-warn{color:#ef5350!important}
  .kpi-ok{color:#26c6da!important}
  .section-title{font-size:1rem;font-weight:600;color:#90caf9;
    border-left:3px solid #1565c0;padding-left:10px;margin:24px 0 10px}
  [data-testid="stTabs"] button{color:#90caf9!important}
  [data-testid="stTabs"] button[aria-selected="true"]{
    border-bottom:2px solid #1565c0!important;color:#e8f0fe!important}
  .stSelectbox label,.stSlider label,.stMultiSelect label,.stRadio label{color:#90caf9!important}
  div[data-baseweb="select"]>div{background:#111827!important;border-color:#1e3a5f!important;color:#e8eaf0!important}
  .stPlotlyChart{border-radius:10px;overflow:hidden}
  hr{border-color:#1e2a45}
</style>
""", unsafe_allow_html=True)

DATA_DIR = "data"

# ── Base layout for all plots
BL = dict(
    paper_bgcolor="#111827", plot_bgcolor="#0f1623",
    font=dict(color="#c8d0e0", family="Inter, Arial"),
    xaxis=dict(gridcolor="#1e2a45", linecolor="#1e3a5f", zerolinecolor="#1e3a5f"),
    yaxis=dict(gridcolor="#1e2a45", linecolor="#1e3a5f", zerolinecolor="#1e3a5f"),
    margin=dict(l=50,r=20,t=50,b=50),
    hoverlabel=dict(bgcolor="#0d1220",font_color="#e8eaf0",bordercolor="#1565c0"),
    legend=dict(bgcolor="rgba(0,0,0,0)",bordercolor="#1e2a45",borderwidth=1),
)
PAL = px.colors.qualitative.Set3

# ══════════════════════════════════════════════════════════════
# PARSERS
# ══════════════════════════════════════════════════════════════
def parse_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    out = {}

    # ── Rodete
    for unit in ["UG1","UG2"]:
        sname = f"{unit}_MED_ALAB_ROD_PER_SAL"
        if sname not in wb.sheetnames: continue
        ws = wb[sname]; rows = []
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            if i<2: continue
            if not row[0] or not isinstance(row[0], datetime): continue
            rows.append({"fecha":row[0].date(),"punto":row[1],
                         **{f"M{j+1}":row[2+j] for j in range(13)}})
        if rows: out[f"rodete_{unit}"] = pd.DataFrame(rows)

    # ── Directrices
    for unit in ["UG1","UG2"]:
        sname = f"{unit}_MED_HOL_ALAB_DIREC"
        if sname not in wb.sheetnames: continue
        ws = wb[sname]; rows = []
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            if i<3: continue
            if not row[0] or not isinstance(row[0], datetime): continue
            if all(v is None for v in row[1:6]): continue
            rows.append({"fecha":row[0].date(),"alabe":row[1],
                         "sup_entrada_A":row[2],"sup_salida_B":row[3],
                         "inf_entrada_A":row[4],"inf_salida_B":row[5]})
        if rows: out[f"directriz_{unit}"] = pd.DataFrame(rows)

    # ── Pista de Freno
    for unit in ["UG1","UG2"]:
        sname = f"{unit}_PLAN_PIST_FRENO"
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        all_rows = list(ws.iter_rows(values_only=True))
        years = []
        for cell in all_rows[0]:
            if cell and isinstance(cell,str):
                m = re.search(r'\b(20\d\d)\b', cell)
                if m: years.append(int(m.group(1)))
        rows = []
        for row in all_rows[4:]:
            if not row[0]: continue
            freno = row[0]
            vals = list(row[1:])
            for i,yr in enumerate(years):
                b = i*4
                v = [vals[b+k] if b+k<len(vals) else None for k in range(4)]
                if all(x is None for x in v): continue
                rows.append({"año":yr,"freno":freno,
                             "DS_interno":v[0],"DS_externo":v[1],
                             "US_interno":v[2],"US_externo":v[3]})
        if rows: out[f"freno_{unit}"] = pd.DataFrame(rows)

    # ── Placa Sello Eje  (usa MED_PLAC_DESGA_SELL_EJE que tiene años como int)
    for unit in ["UG1","UG2"]:
        sname = f"{unit}_MED_PLAC_DESGA_SELL_EJE"
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        all_rows = list(ws.iter_rows(values_only=True))
        years = [int(c) for c in all_rows[1][1:] if c and str(c).strip() and
                 re.match(r'20\d\d', str(c))]
        rows = []
        for row in all_rows[2:]:
            if not row[0]: continue
            punto = str(row[0]).strip()
            vals = list(row[1:])
            for i,yr in enumerate(years):
                b = i*2
                v1 = vals[b]   if b   < len(vals) else None
                v2 = vals[b+1] if b+1 < len(vals) else None
                if v1 is None and v2 is None: continue
                rows.append({"año":yr,"punto":punto,"sensor_1":v1,"sensor_2":v2})
        if rows: out[f"sello_{unit}"] = pd.DataFrame(rows)

    # ── Cojinete Guía
    for unit in ["UG1","UG2"]:
        sname = f"{unit}_MED_HOLG_CAS_COJ_GUIA"
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        all_rows = list(ws.iter_rows(values_only=True))
        years = [int(c) for c in all_rows[1][1:] if c and isinstance(c,(int,float))]
        rows = []
        for row in all_rows[2:]:
            if not row[0]: continue
            for i,yr in enumerate(years):
                v = row[1+i] if 1+i < len(row) else None
                rows.append({"año":yr,"punto":str(row[0]),"holgura":v})
        if rows: out[f"cojinete_{unit}"] = pd.DataFrame(rows)

    wb.close()
    return out

# ══════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════
@st.cache_data
def load_csv_data():
    data = {}
    prefixes = ["rodete","directriz","freno","sello","cojinete"]
    units    = ["UG1","UG2"]
    for p in prefixes:
        for u in units:
            key  = f"{p}_{u}"
            path = os.path.join(DATA_DIR, f"data_{key}.csv")
            if os.path.exists(path):
                df = pd.read_csv(path)
                if "fecha" in df.columns:
                    df["fecha"] = pd.to_datetime(df["fecha"])
                data[key] = df
    return data

def get_data():
    if "live_data" in st.session_state:
        return st.session_state["live_data"]
    return load_csv_data()

# ══════════════════════════════════════════════════════════════
# REGRESSION HELPERS
# ══════════════════════════════════════════════════════════════
def year_to_date(years_series):
    """Convert year integers to datetime for x-axis."""
    return pd.to_datetime(years_series.astype(int).astype(str) + "-01-01")

def days_from_origin(dates):
    t0 = dates.min()
    return (dates - t0).dt.days.values, t0

def _r2(y,yp):
    ss=np.sum((y-np.mean(y))**2)
    return 1-np.sum((y-yp)**2)/ss if ss>0 else 0

def _rmse(y,yp): return np.sqrt(np.mean((y-yp)**2))

def fit_linear(x,y):
    s,b,*_=stats.linregress(x,y); yp=s*x+b
    return {"name":"Lineal","pred":lambda xx,s=s,b=b:s*xx+b,
            "r2":_r2(y,yp),"rmse":_rmse(y,yp),"color":"#42a5f5"}

def fit_poly(x,y,deg):
    c=np.polyfit(x,y,deg); p=np.poly1d(c); yp=p(x)
    return {"name":f"Polinómica g{deg}","pred":lambda xx,p=p:p(xx),
            "r2":_r2(y,yp),"rmse":_rmse(y,yp),
            "color":"#ab47bc" if deg==2 else "#ce93d8"}

def fit_exp(x,y):
    try:
        def fn(xx,a,b,c): return a*np.exp(b*xx)+c
        popt,_=curve_fit(fn,x,y,p0=[y.max()-y.min()+.01,-1e-4,y.min()],maxfev=6000)
        yp=fn(x,*popt)
        return {"name":"Exponencial","pred":lambda xx,p=popt:fn(xx,*p),
                "r2":_r2(y,yp),"rmse":_rmse(y,yp),"color":"#ef5350"}
    except: return None

def fit_pow(x,y):
    try:
        def fn(xx,a,b,c): return a*(np.abs(np.where(xx==0,1,xx))**b)+c
        popt,_=curve_fit(fn,x,y,p0=[1,.5,y.min()],maxfev=6000)
        yp=fn(x,*popt)
        return {"name":"Potencial","pred":lambda xx,p=popt:fn(xx,*p),
                "r2":_r2(y,yp),"rmse":_rmse(y,yp),"color":"#ff7043"}
    except: return None

def get_all_fits(x,y):
    fits=[fit_linear(x,y),fit_poly(x,y,2),fit_poly(x,y,3)]
    for fn in [fit_exp,fit_pow]:
        f=fn(x,y)
        if f: fits.append(f)
    return sorted(fits,key=lambda f:-f["r2"])

def forecast_crossing(fit,x_max,y_target,t0,horizon_days):
    xs=np.linspace(0,x_max+horizon_days,3000)
    ys=fit["pred"](xs)
    idx=np.where(np.diff(np.sign(ys-y_target)))[0]
    if len(idx): return t0+timedelta(days=int(xs[idx[0]]))
    return None

def color_r2(v):
    if v>=0.95: return "background:#1b5e20;color:#a5d6a7"
    if v>=0.80: return "background:#0d2b4e;color:#90caf9"
    return "background:#3e1e1e;color:#ef9a9a"

def add_today(fig):
    today_str=datetime.today().strftime("%Y-%m-%d")
    fig.add_shape(type="line",x0=today_str,x1=today_str,y0=0,y1=1,
                  xref="x",yref="paper",line=dict(color="#546e8a",dash="dot",width=1.5))
    fig.add_annotation(x=today_str,y=1,xref="x",yref="paper",text="Hoy",
                       showarrow=False,font=dict(color="#546e8a",size=11),
                       xanchor="left",yanchor="top")

def regression_block(fig, x, t0, y, sel_fits, threshold, forecast_yrs):
    horizon=int(forecast_yrs*365)
    x_ext=np.linspace(0,x.max()+horizon,500)
    d_ext=[(t0+timedelta(days=int(d))).strftime("%Y-%m-%d") for d in x_ext]
    forecasts=[]
    for fit in sel_fits:
        yf=fit["pred"](x_ext)
        r,g,b=int(fit["color"][1:3],16),int(fit["color"][3:5],16),int(fit["color"][5:7],16)
        fig.add_trace(go.Scatter(x=d_ext+d_ext[::-1],
            y=list(yf+fit["rmse"])+list((yf-fit["rmse"])[::-1]),
            fill="toself",line=dict(color="rgba(0,0,0,0)"),
            fillcolor=f"rgba({r},{g},{b},0.10)",showlegend=False,hoverinfo="skip"))
        fig.add_trace(go.Scatter(x=d_ext,y=yf,mode="lines",
            name=f"{fit['name']}  R²={fit['r2']:.4f}",
            line=dict(color=fit["color"],width=2.5),
            hovertemplate=f"<b>{fit['name']}</b><br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
        if threshold>0:
            fd=forecast_crossing(fit,x.max(),threshold,t0,horizon)
            if fd:
                forecasts.append({"Modelo":fit["name"],
                                   "Fecha estimada":fd.strftime("%Y-%m-%d"),
                                   "R²":round(fit["r2"],4),"RMSE":round(fit["rmse"],4)})
    if threshold>0:
        fig.add_hline(y=threshold,line_dash="dot",line_color="#ef5350",
                      annotation_text=f"Límite {threshold} mm",annotation_font_color="#ef5350")
    add_today(fig)
    return sorted(forecasts,key=lambda x:-x["R²"])

def show_forecasts(forecasts,threshold,forecast_yr):
    if forecasts:
        st.markdown('<div class="section-title">📅 Fechas estimadas de falla</div>',
                    unsafe_allow_html=True)
        st.dataframe(pd.DataFrame(forecasts),use_container_width=True)
        bf=forecasts[0]
        dl=(pd.to_datetime(bf["Fecha estimada"])-pd.Timestamp.today()).days
        msg=(f"Según **{bf['Modelo']}** (R²={bf['R²']:.4f}), se estima alcanzar "
             f"**{threshold} mm** el **{bf['Fecha estimada']}**")
        if dl>0:
            st.info(f"⏱️ {msg} — en ~**{dl} días** desde hoy.")
        else:
            st.warning(f"⚠️ {msg} — posiblemente ya alcanzado.")
    elif threshold>0:
        st.info(f"ℹ️ No se proyecta alcanzar {threshold} mm en {forecast_yr} años.")

def kpi_row(kpis):
    st.markdown('<div class="kpi-grid">',unsafe_allow_html=True)
    for col_st,(label,val,sub,cls) in zip(st.columns(len(kpis)),kpis):
        with col_st:
            st.markdown(f'<div class="kpi-card"><div class="kpi-label">{label}</div>'
                        f'<div class="kpi-value {cls}">{val}</div>'
                        f'<div class="kpi-sub">{sub}</div></div>',unsafe_allow_html=True)
    st.markdown("</div>",unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# SIDEBAR DATA PANEL
# ══════════════════════════════════════════════════════════════
def sidebar_data_panel():
    st.sidebar.markdown("### 📂 Actualizar datos")
    st.sidebar.markdown(
        '<div style="background:#0d1a2e;border:1px dashed #1565c0;border-radius:8px;'
        'padding:12px;margin-bottom:10px">'
        '<p style="color:#7986cb;font-size:.78rem;margin:0">'
        '① Sube el Excel → ② Procesa → ③ Descarga ZIP → ④ Sube /data a GitHub</p></div>',
        unsafe_allow_html=True)
    uploaded=st.sidebar.file_uploader("Subir .xlsm / .xlsx",type=["xlsm","xlsx"])
    if uploaded is not None:
        with st.sidebar:
            with st.spinner("Procesando…"):
                try:
                    dfs=parse_excel(uploaded)
                    live={}
                    for k,df in dfs.items():
                        df2=pd.DataFrame(df)
                        if "fecha" in df2.columns:
                            df2["fecha"]=pd.to_datetime(df2["fecha"])
                        live[k]=df2
                    st.session_state["live_data"]=live
                    st.session_state["csv_ready"]=dfs
                    st.success(f"✅ {sum(len(d) for d in dfs.values())} registros")
                except Exception as e:
                    st.error(f"Error: {e}")
    if "csv_ready" in st.session_state:
        st.sidebar.markdown("---")
        buf=io.BytesIO()
        with zipfile.ZipFile(buf,"w",zipfile.ZIP_DEFLATED) as zf:
            for key,df in st.session_state["csv_ready"].items():
                zf.writestr(f"data/data_{key}.csv",pd.DataFrame(df).to_csv(index=False).encode())
        buf.seek(0)
        st.sidebar.download_button("⬇️ data_vernova.zip",data=buf,
                                    file_name="data_vernova.zip",mime="application/zip")
    st.sidebar.markdown("---")

# ══════════════════════════════════════════════════════════════
# MÓDULO 1 — RODETE
# ══════════════════════════════════════════════════════════════
def mod_rodete(data):
    keys={k:v for k,v in data.items() if k.startswith("rodete_")}
    if not keys:
        st.warning("No hay datos de Rodete disponibles."); return

    unit_map={k.replace("rodete_",""):k for k in keys}
    unit=st.sidebar.selectbox("🔧 Unidad",list(unit_map.keys()),key="rod_unit")
    df_unit=keys[unit_map[unit]]
    puntos=sorted(df_unit["punto"].unique())
    punto=st.sidebar.selectbox("📍 Zona (D1–D5)",puntos,key="rod_punto")
    df_p=df_unit[df_unit["punto"]==punto].sort_values("fecha").reset_index(drop=True)
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📊 Regresión")
    reg_opts=["Lineal","Polinómica g2","Polinómica g3","Exponencial","Potencial"]
    reg_sel=st.sidebar.multiselect("Modelos",reg_opts,
                                    default=["Lineal","Polinómica g2","Exponencial"],key="rod_reg")
    med_reg=st.sidebar.slider("Posición para regresión",1,13,7,key="rod_med")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚠️ Pronóstico")
    threshold=st.sidebar.number_input("Espesor límite crítico (mm)",0.0,30.0,14.0,0.1,key="rod_thr")
    forecast_yr=st.sidebar.slider("Horizonte (años)",1,10,4,key="rod_fc")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🔍 Vista")
    med_vis=st.sidebar.multiselect("Posiciones a graficar",list(range(1,14)),
                                    default=list(range(1,14)),key="rod_vis")

    m_cols=[f"M{i}" for i in range(1,14)]
    r_first=df_p.iloc[0]; r_last=df_p.iloc[-1]
    avg_f=np.nanmean([r_first.get(c,np.nan) for c in m_cols])
    avg_l=np.nanmean([r_last.get(c,np.nan) for c in m_cols])
    wear=avg_f-avg_l
    n_days=(df_p["fecha"].max()-df_p["fecha"].min()).days
    rate_yr=wear/n_days*365 if n_days>0 else 0

    kpi_row([
        ("Unidad",unit,"",""),
        ("Zona",punto,"",""),
        ("Puntos de muestreo",len(df_p),"registros","kpi-ok"),
        ("Espesor inicial",f"{avg_f:.3f} mm",df_p['fecha'].min().strftime('%Y-%m-%d'),""),
        ("Espesor actual",f"{avg_l:.3f} mm",df_p['fecha'].max().strftime('%Y-%m-%d'),
         "kpi-warn" if avg_l<threshold else "kpi-ok"),
        ("Desgaste acumulado",f"{wear:.3f} mm","primera → última medición",
         "kpi-warn" if wear>1 else ""),
        ("Tasa anual",f"{rate_yr:.3f} mm/a","promedio histórico",
         "kpi-warn" if rate_yr>0.3 else ""),
    ])

    tab1,tab2,tab3,tab4=st.tabs(["📏 Evolución temporal","📐 Perfil espacial",
                                  "📈 Regresión & Pronóstico","🌡️ Mapa de calor & Desgaste"])
    with tab1:
        st.markdown('<div class="section-title">Evolución del espesor por posición a lo largo del tiempo</div>',
                    unsafe_allow_html=True)
        if med_vis:
            fig=go.Figure()
            for i,m in enumerate(med_vis):
                col=f"M{m}"
                if col not in df_p.columns: continue
                fig.add_trace(go.Scatter(x=df_p["fecha"].dt.strftime("%Y-%m-%d"),y=df_p[col],
                    mode="lines+markers",name=f"Pos {m}",
                    line=dict(color=PAL[i%len(PAL)],width=2),marker=dict(size=7),
                    hovertemplate=f"Pos {m}<br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
            fig.update_layout(**BL,height=400,
                title=dict(text=f"📏 Evolución — {unit} · {punto}",font_size=14,x=0.01),
                xaxis_title="Fecha",yaxis_title="Espesor (mm)")
            st.plotly_chart(fig,use_container_width=True)
        else:
            st.info("Selecciona al menos una posición.")
        st.markdown('<div class="section-title">Datos registrados</div>',unsafe_allow_html=True)
        disp=df_p.copy(); disp["fecha"]=disp["fecha"].dt.strftime("%Y-%m-%d")
        disp.columns=["Fecha","Zona"]+[f"Pos {i}" for i in range(1,14)]
        st.dataframe(disp.style.background_gradient(cmap="RdYlGn",
                     subset=[f"Pos {i}" for i in range(1,14)]),
                     use_container_width=True,height=220)

    with tab2:
        st.markdown('<div class="section-title">Perfil espacial por punto de muestreo</div>',
                    unsafe_allow_html=True)
        fig2=go.Figure()
        clrs=["#42a5f5","#26c6da","#66bb6a","#ffa726","#ef5350"]
        for i,(_,row) in enumerate(df_p.iterrows()):
            vals=[row.get(c,np.nan) for c in m_cols]
            fig2.add_trace(go.Scatter(x=[str(j) for j in range(1,14)],y=vals,
                mode="lines+markers",name=row["fecha"].strftime("%Y-%m-%d"),
                line=dict(color=clrs[i%len(clrs)],width=2),marker=dict(size=8),
                fill="tozeroy" if i==0 else "none",fillcolor="rgba(66,165,245,0.06)",
                hovertemplate="Pos %{x}: %{y:.3f} mm<extra></extra>"))
        fig2.update_layout(**BL,height=380,
            title=dict(text=f"📐 Perfil espacial — {unit} · {punto}",font_size=14,x=0.01),
            xaxis_title="Posición (1→13)",yaxis_title="Espesor (mm)")
        st.plotly_chart(fig2,use_container_width=True)
        c1,c2=st.columns(2)
        for cs,rd,lbl,clr in [
            (c1,r_first,f"Primera — {df_p['fecha'].min().strftime('%Y-%m-%d')}","#42a5f5"),
            (c2,r_last, f"Última  — {df_p['fecha'].max().strftime('%Y-%m-%d')}","#ef5350")]:
            with cs:
                st.markdown(f"**{lbl}**")
                vals=[rd.get(c,np.nan) for c in m_cols]
                fb=go.Figure(go.Bar(x=[str(i) for i in range(1,14)],y=vals,
                    marker_color=clr,hovertemplate="Pos %{x}: %{y:.3f} mm<extra></extra>"))
                fb.update_layout(**BL,height=280,yaxis_title="mm",xaxis_title="Posición",
                    title=dict(text=lbl,font_size=11))
                st.plotly_chart(fb,use_container_width=True)

    with tab3:
        col=f"M{med_reg}"
        st.markdown(f'<div class="section-title">Regresión — {unit} · {punto} · Posición {med_reg}</div>',
                    unsafe_allow_html=True)
        if col in df_p.columns:
            df_s=df_p.sort_values("fecha")
            x,t0=days_from_origin(df_s["fecha"])
            y=df_s[col].values.astype(float)
            all_fits=get_all_fits(x,y)
            sel_fits=[f for f in all_fits if f["name"] in reg_sel] if reg_sel else all_fits
            fig3=go.Figure()
            fig3.add_trace(go.Scatter(x=df_s["fecha"].dt.strftime("%Y-%m-%d"),y=y,
                mode="markers",name="Medición real",
                marker=dict(size=11,color="#ffd54f",symbol="circle",
                            line=dict(color="#ff8f00",width=1.5)),
                hovertemplate="Fecha: %{x}<br>%{y:.3f} mm<extra></extra>"))
            forecasts=regression_block(fig3,x,t0,y,sel_fits,threshold,forecast_yr)
            fig3.update_layout(**BL,height=460,
                title=dict(text=f"📈 Regresión — {unit}·{punto} · Pos {med_reg}",font_size=14,x=0.01),
                xaxis_title="Fecha",yaxis_title="Espesor (mm)")
            st.plotly_chart(fig3,use_container_width=True)
            show_forecasts(forecasts,threshold,forecast_yr)

        st.markdown('<div class="section-title">Mejor modelo por posición</div>',unsafe_allow_html=True)
        x_d,_=days_from_origin(df_p.sort_values("fecha")["fecha"])
        rows_tbl=[]
        for i in range(1,14):
            c2=f"M{i}"
            if c2 not in df_p.columns: continue
            yc=df_p.sort_values("fecha")[c2].values.astype(float)
            if np.any(np.isnan(yc)): continue
            best=get_all_fits(x_d,yc)[0]
            rows_tbl.append({"Posición":i,"Mejor modelo":best["name"],
                              "R²":round(best["r2"],4),"RMSE":round(best["rmse"],4),
                              "Espesor actual (mm)":round(float(df_p.sort_values("fecha")[c2].iloc[-1]),3)})
        if rows_tbl:
            st.dataframe(pd.DataFrame(rows_tbl).style.map(color_r2,subset=["R²"]),
                         use_container_width=True,height=420)

    with tab4:
        last=df_unit["fecha"].max(); df_l=df_unit[df_unit["fecha"]==last]
        puntos_all=sorted(df_l["punto"].unique())
        z=[[df_l[df_l["punto"]==p].iloc[0].get(c,np.nan) for c in m_cols] for p in puntos_all]
        fig_hm=go.Figure(go.Heatmap(z=z,x=[str(i) for i in range(1,14)],y=puntos_all,
            colorscale="RdYlGn_r",text=[[f"{v:.2f}" for v in row] for row in z],
            texttemplate="%{text}",textfont=dict(size=10,color="#000"),
            hovertemplate="Zona:%{y} Pos:%{x}<br>%{z:.3f} mm<extra></extra>",
            colorbar=dict(title=dict(text="mm",font=dict(color="#c8d0e0")),tickfont=dict(color="#c8d0e0"))))
        fig_hm.update_layout(**BL,height=320,
            title=dict(text=f"🌡️ Mapa de calor — {unit} · {pd.Timestamp(last).strftime('%Y-%m-%d')}",
                       font_size=14,x=0.01),xaxis_title="Posición (1→13)",yaxis_title="Zona")
        st.markdown('<div class="section-title">Mapa de calor — última medición</div>',unsafe_allow_html=True)
        st.plotly_chart(fig_hm,use_container_width=True)

        dates_all=sorted(df_unit["fecha"].unique())
        if len(dates_all)>=2:
            df_f2=df_unit[df_unit["fecha"]==dates_all[0]]
            df_l2=df_unit[df_unit["fecha"]==dates_all[-1]]
            fig_d=go.Figure()
            clrs2=["#42a5f5","#26c6da","#66bb6a","#ffa726","#ef5350"]
            for i,p in enumerate(sorted(df_unit["punto"].unique())):
                rf=df_f2[df_f2["punto"]==p]; rl=df_l2[df_l2["punto"]==p]
                if rf.empty or rl.empty: continue
                delta=[rf.iloc[0].get(c,np.nan)-rl.iloc[0].get(c,np.nan) for c in m_cols]
                fig_d.add_trace(go.Bar(x=[str(j) for j in range(1,14)],y=delta,name=p,
                    marker_color=clrs2[i%len(clrs2)],
                    hovertemplate=f"Zona {p} Pos %{{x}}<br>Δ=%{{y:.3f}} mm<extra></extra>"))
            fig_d.update_layout(**BL,barmode="group",height=360,
                title=dict(text=f"📉 Desgaste acumulado — {unit} "
                               f"({pd.Timestamp(dates_all[0]).strftime('%Y-%m-%d')} → "
                               f"{pd.Timestamp(dates_all[-1]).strftime('%Y-%m-%d')})",
                           font_size=14,x=0.01),
                xaxis_title="Posición",yaxis_title="Reducción de espesor (mm)")
            st.markdown('<div class="section-title">Desgaste acumulado total</div>',unsafe_allow_html=True)
            st.plotly_chart(fig_d,use_container_width=True)

        st.markdown('<div class="section-title">Comparativo entre unidades</div>',unsafe_allow_html=True)
        comp_rows=[]
        for k,df_u in keys.items():
            ul=k.replace("rodete_","")
            for p in sorted(df_u["punto"].unique()):
                for _,row in df_u[df_u["punto"]==p].iterrows():
                    comp_rows.append({"Unidad":ul,"Zona":p,"Fecha":row["fecha"],
                                      "Avg":np.nanmean([row.get(c,np.nan) for c in m_cols])})
        if comp_rows:
            comp_df=pd.DataFrame(comp_rows)
            fig_cmp=go.Figure()
            ug_c={"UG1":"#42a5f5","UG2":"#ef5350"}
            dashes=["solid","dash","dot","dashdot","longdash"]
            for ug in comp_df["Unidad"].unique():
                for j,p in enumerate(sorted(comp_df["Zona"].unique())):
                    sub=comp_df[(comp_df["Unidad"]==ug)&(comp_df["Zona"]==p)].sort_values("Fecha")
                    if sub.empty: continue
                    fig_cmp.add_trace(go.Scatter(
                        x=sub["Fecha"].dt.strftime("%Y-%m-%d"),y=sub["Avg"],
                        mode="lines+markers",name=f"{ug}·{p}",
                        line=dict(color=ug_c.get(ug,"#90caf9"),width=2,dash=dashes[j%len(dashes)]),
                        marker=dict(size=7),
                        hovertemplate=f"{ug} {p}<br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
            fig_cmp.update_layout(**BL,height=420,
                title=dict(text="Espesor promedio por unidad y zona",font_size=14,x=0.01),
                xaxis_title="Fecha",yaxis_title="Espesor promedio (mm)")
            st.plotly_chart(fig_cmp,use_container_width=True)

# ══════════════════════════════════════════════════════════════
# MÓDULO 2 — DIRECTRICES
# ══════════════════════════════════════════════════════════════
DIR_COLS=["sup_entrada_A","sup_salida_B","inf_entrada_A","inf_salida_B"]
DIR_LABELS={"sup_entrada_A":"Sup. Entrada A","sup_salida_B":"Sup. Salida B",
             "inf_entrada_A":"Inf. Entrada A","inf_salida_B":"Inf. Salida B"}
DIR_COLORS={"sup_entrada_A":"#42a5f5","sup_salida_B":"#26c6da",
             "inf_entrada_A":"#66bb6a","inf_salida_B":"#ffa726"}

def mod_directriz(data):
    keys={k:v for k,v in data.items() if k.startswith("directriz_")}
    if not keys:
        st.warning("No hay datos de Directrices disponibles."); return

    unit_map={k.replace("directriz_",""):k for k in keys}
    unit=st.sidebar.selectbox("🔧 Unidad",list(unit_map.keys()),key="dir_unit")
    df_unit=keys[unit_map[unit]]
    alabes=sorted(df_unit["alabe"].unique())
    alabe=st.sidebar.selectbox("🔢 Álabe (1–20)",alabes,key="dir_alabe")
    df_a=df_unit[df_unit["alabe"]==alabe].sort_values("fecha").reset_index(drop=True)
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📊 Regresión")
    lbl2col={v:k for k,v in DIR_LABELS.items()}
    punto_reg=st.sidebar.selectbox("Punto de medida",list(DIR_LABELS.values()),key="dir_preg")
    col_reg=lbl2col[punto_reg]
    reg_opts=["Lineal","Polinómica g2","Polinómica g3","Exponencial","Potencial"]
    reg_sel=st.sidebar.multiselect("Modelos",reg_opts,
                                    default=["Lineal","Polinómica g2","Exponencial"],key="dir_reg")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚠️ Pronóstico")
    threshold=st.sidebar.number_input("Holgura límite crítica (mm)",0.0,5.0,0.15,0.01,key="dir_thr")
    forecast_yr=st.sidebar.slider("Horizonte (años)",1,10,4,key="dir_fc")

    r_first=df_a.iloc[0]; r_last=df_a.iloc[-1]
    avg_f=np.nanmean([r_first.get(c,np.nan) for c in DIR_COLS])
    avg_l=np.nanmean([r_last.get(c,np.nan) for c in DIR_COLS])
    wear=avg_f-avg_l
    n_days=(df_a["fecha"].max()-df_a["fecha"].min()).days
    rate_yr=wear/n_days*365 if n_days>0 else 0

    kpi_row([
        ("Unidad",unit,"",""),
        ("Álabe",f"#{alabe}","",""),
        ("Puntos de muestreo",len(df_a),"registros","kpi-ok"),
        ("Holgura inicial prom",f"{avg_f:.3f} mm",df_a['fecha'].min().strftime('%Y-%m-%d'),""),
        ("Holgura actual prom",f"{avg_l:.3f} mm",df_a['fecha'].max().strftime('%Y-%m-%d'),
         "kpi-warn" if avg_l<threshold else "kpi-ok"),
        ("Variación acumulada",f"{wear:.3f} mm","primera → última medición",
         "kpi-warn" if abs(wear)>0.1 else ""),
        ("Tasa anual",f"{rate_yr:.3f} mm/a","promedio histórico",
         "kpi-warn" if abs(rate_yr)>0.05 else ""),
    ])

    tab1,tab2,tab3,tab4=st.tabs(["📏 Evolución temporal","📐 Perfil por álabe",
                                  "📈 Regresión & Pronóstico","🌡️ Mapa de calor"])
    with tab1:
        st.markdown('<div class="section-title">Evolución de holguras en los 4 puntos de medida</div>',
                    unsafe_allow_html=True)
        fig1=go.Figure()
        for col,lbl in DIR_LABELS.items():
            if col not in df_a.columns: continue
            fig1.add_trace(go.Scatter(x=df_a["fecha"].dt.strftime("%Y-%m-%d"),y=df_a[col],
                mode="lines+markers",name=lbl,
                line=dict(color=DIR_COLORS[col],width=2),marker=dict(size=8),
                hovertemplate=f"{lbl}<br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
        fig1.update_layout(**BL,height=400,
            title=dict(text=f"📏 Holguras — {unit} · Álabe #{alabe}",font_size=14,x=0.01),
            xaxis_title="Fecha",yaxis_title="Holgura (mm)")
        st.plotly_chart(fig1,use_container_width=True)
        disp=df_a.copy(); disp["fecha"]=disp["fecha"].dt.strftime("%Y-%m-%d")
        disp=disp[["fecha","alabe"]+DIR_COLS]
        disp.columns=["Fecha","Álabe"]+list(DIR_LABELS.values())
        st.dataframe(disp.style.background_gradient(cmap="RdYlGn_r",
                     subset=list(DIR_LABELS.values())),use_container_width=True,height=260)

    with tab2:
        st.markdown('<div class="section-title">Todos los álabes — un punto de medida</div>',
                    unsafe_allow_html=True)
        punto_vis=st.selectbox("Ver punto",list(DIR_LABELS.values()),key="dir_tab2")
        col_vis=lbl2col[punto_vis]
        fechas_all=sorted(df_unit["fecha"].unique())
        clrs=["#42a5f5","#26c6da","#66bb6a","#ffa726","#ef5350","#ab47bc","#ff7043"]
        fig2=go.Figure()
        for i,f in enumerate(fechas_all):
            df_f=df_unit[df_unit["fecha"]==f].sort_values("alabe")
            fig2.add_trace(go.Scatter(x=df_f["alabe"].astype(str),y=df_f[col_vis],
                mode="lines+markers",name=pd.Timestamp(f).strftime("%Y-%m-%d"),
                line=dict(color=clrs[i%len(clrs)],width=2),marker=dict(size=7),
                hovertemplate="Álabe %{x}<br>%{y:.3f} mm<extra></extra>"))
        fig2.update_layout(**BL,height=400,
            title=dict(text=f"📐 {punto_vis} — todos los álabes — {unit}",font_size=14,x=0.01),
            xaxis_title="N° Álabe",yaxis_title="Holgura (mm)")
        st.plotly_chart(fig2,use_container_width=True)
        c1,c2=st.columns(2)
        for cs,fecha,lbl,clr in [
            (c1,fechas_all[0], f"Primera — {pd.Timestamp(fechas_all[0]).strftime('%Y-%m-%d')}","#42a5f5"),
            (c2,fechas_all[-1],f"Última  — {pd.Timestamp(fechas_all[-1]).strftime('%Y-%m-%d')}","#ef5350")]:
            with cs:
                st.markdown(f"**{lbl}**")
                df_snap=df_unit[df_unit["fecha"]==fecha].sort_values("alabe")
                fb=go.Figure()
                for col,l in DIR_LABELS.items():
                    fb.add_trace(go.Bar(x=df_snap["alabe"].astype(str),y=df_snap[col],
                        name=l,marker_color=DIR_COLORS[col],
                        hovertemplate=f"{l} Álabe %{{x}}: %{{y:.3f}} mm<extra></extra>"))
                fb.update_layout(**BL,barmode="group",height=300,
                    yaxis_title="mm",xaxis_title="N° Álabe",title=dict(text=lbl,font_size=11))
                st.plotly_chart(fb,use_container_width=True)

    with tab3:
        st.markdown(f'<div class="section-title">Regresión — {unit} · Álabe #{alabe} · {punto_reg}</div>',
                    unsafe_allow_html=True)
        if col_reg in df_a.columns:
            df_s=df_a.sort_values("fecha")
            x,t0=days_from_origin(df_s["fecha"])
            y=df_s[col_reg].values.astype(float)
            all_fits=get_all_fits(x,y)
            sel_fits=[f for f in all_fits if f["name"] in reg_sel] if reg_sel else all_fits
            fig3=go.Figure()
            fig3.add_trace(go.Scatter(x=df_s["fecha"].dt.strftime("%Y-%m-%d"),y=y,
                mode="markers",name="Medición real",
                marker=dict(size=11,color="#ffd54f",symbol="circle",
                            line=dict(color="#ff8f00",width=1.5)),
                hovertemplate="Fecha: %{x}<br>%{y:.3f} mm<extra></extra>"))
            forecasts=regression_block(fig3,x,t0,y,sel_fits,threshold,forecast_yr)
            fig3.update_layout(**BL,height=460,
                title=dict(text=f"📈 Regresión — {unit} · Álabe #{alabe} · {punto_reg}",
                           font_size=14,x=0.01),xaxis_title="Fecha",yaxis_title="Holgura (mm)")
            st.plotly_chart(fig3,use_container_width=True)
            show_forecasts(forecasts,threshold,forecast_yr)

        st.markdown('<div class="section-title">Mejor modelo por punto de medida</div>',unsafe_allow_html=True)
        x_d,_=days_from_origin(df_a.sort_values("fecha")["fecha"])
        rows_tbl=[]
        for col,lbl in DIR_LABELS.items():
            if col not in df_a.columns: continue
            yc=df_a.sort_values("fecha")[col].values.astype(float)
            if np.any(np.isnan(yc)): continue
            best=get_all_fits(x_d,yc)[0]
            rows_tbl.append({"Punto":lbl,"Mejor modelo":best["name"],
                              "R²":round(best["r2"],4),"RMSE":round(best["rmse"],4),
                              "Holgura actual (mm)":round(float(df_a.sort_values("fecha")[col].iloc[-1]),3)})
        if rows_tbl:
            st.dataframe(pd.DataFrame(rows_tbl).style.map(color_r2,subset=["R²"]),
                         use_container_width=True,height=220)

    with tab4:
        st.markdown('<div class="section-title">Mapa de calor — última medición (álabes × puntos)</div>',
                    unsafe_allow_html=True)
        last=df_unit["fecha"].max()
        df_last=df_unit[df_unit["fecha"]==last].sort_values("alabe")
        z_hm=[[row.get(c,np.nan) for c in DIR_COLS] for _,row in df_last.iterrows()]
        fig_hm=go.Figure(go.Heatmap(z=z_hm,x=list(DIR_LABELS.values()),
            y=df_last["alabe"].astype(str).tolist(),colorscale="RdYlGn_r",
            text=[[f"{v:.3f}" for v in row] for row in z_hm],
            texttemplate="%{text}",textfont=dict(size=9,color="#000"),
            hovertemplate="Álabe %{y} · %{x}<br>%{z:.3f} mm<extra></extra>",
            colorbar=dict(title=dict(text="mm",font=dict(color="#c8d0e0")),tickfont=dict(color="#c8d0e0"))))
        fig_hm.update_layout(**BL,height=520,
            title=dict(text=f"🌡️ Mapa de calor — {unit} · {pd.Timestamp(last).strftime('%Y-%m-%d')}",
                       font_size=14,x=0.01),xaxis_title="Punto",yaxis_title="N° Álabe")
        st.plotly_chart(fig_hm,use_container_width=True)

        st.markdown('<div class="section-title">Comparativo entre unidades — holgura promedio</div>',
                    unsafe_allow_html=True)
        comp_rows=[]
        for k,df_u in keys.items():
            ul=k.replace("directriz_","")
            for _,row in df_u.iterrows():
                comp_rows.append({"Unidad":ul,"Fecha":row["fecha"],
                                   "Avg":np.nanmean([row.get(c,np.nan) for c in DIR_COLS])})
        if comp_rows:
            comp_df=pd.DataFrame(comp_rows)
            fig_cmp=go.Figure()
            ug_c={"UG1":"#42a5f5","UG2":"#ef5350"}
            for ug in comp_df["Unidad"].unique():
                sub=comp_df[comp_df["Unidad"]==ug].groupby("Fecha")["Avg"].mean().reset_index()
                fig_cmp.add_trace(go.Scatter(x=sub["Fecha"].dt.strftime("%Y-%m-%d"),y=sub["Avg"],
                    mode="lines+markers",name=ug,
                    line=dict(color=ug_c.get(ug,"#90caf9"),width=2),marker=dict(size=7),
                    hovertemplate=f"{ug}<br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
            fig_cmp.update_layout(**BL,height=360,
                title=dict(text="Holgura promedio por unidad",font_size=14,x=0.01),
                xaxis_title="Fecha",yaxis_title="Holgura promedio (mm)")
            st.plotly_chart(fig_cmp,use_container_width=True)

# ══════════════════════════════════════════════════════════════
# MÓDULO 3 — PISTA DE FRENO
# ══════════════════════════════════════════════════════════════
FRENO_COLS=["DS_interno","DS_externo","US_interno","US_externo"]
FRENO_LABELS={"DS_interno":"Aguas Abajo Interno","DS_externo":"Aguas Abajo Externo",
               "US_interno":"Aguas Arriba Interno","US_externo":"Aguas Arriba Externo"}
FRENO_COLORS={"DS_interno":"#42a5f5","DS_externo":"#26c6da",
               "US_interno":"#66bb6a","US_externo":"#ffa726"}
FRENO_REF=5.0  # mm nominal

def mod_freno(data):
    keys={k:v for k,v in data.items() if k.startswith("freno_")}
    if not keys:
        st.warning("No hay datos de Pista de Freno disponibles."); return

    unit_map={k.replace("freno_",""):k for k in keys}
    unit=st.sidebar.selectbox("🔧 Unidad",list(unit_map.keys()),key="freno_unit")
    df_unit=keys[unit_map[unit]]
    frenos=sorted(df_unit["freno"].unique())
    freno=st.sidebar.selectbox("🛑 Freno",frenos,key="freno_sel")
    df_f=df_unit[df_unit["freno"]==freno].sort_values("año").reset_index(drop=True)
    df_f["fecha"]=pd.to_datetime(df_f["año"].astype(str)+"-01-01")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📊 Regresión")
    lbl2col={v:k for k,v in FRENO_LABELS.items()}
    punto_reg=st.sidebar.selectbox("Punto de medida",list(FRENO_LABELS.values()),key="freno_preg")
    col_reg=lbl2col[punto_reg]
    reg_opts=["Lineal","Polinómica g2","Polinómica g3","Exponencial","Potencial"]
    reg_sel=st.sidebar.multiselect("Modelos",reg_opts,
                                    default=["Lineal","Polinómica g2","Exponencial"],key="freno_reg")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚠️ Pronóstico")
    threshold=st.sidebar.number_input("Planicidad límite (mm)",0.0,5.0,FRENO_REF,0.1,key="freno_thr")
    forecast_yr=st.sidebar.slider("Horizonte (años)",1,10,4,key="freno_fc")

    r_first=df_f.iloc[0]; r_last=df_f.iloc[-1]
    avg_f_val=np.nanmean([r_first.get(c,np.nan) for c in FRENO_COLS])
    avg_l_val=np.nanmean([r_last.get(c,np.nan) for c in FRENO_COLS])
    wear=avg_f_val-avg_l_val
    n_yrs=r_last["año"]-r_first["año"]
    rate_yr=wear/n_yrs if n_yrs>0 else 0

    kpi_row([
        ("Unidad",unit,"",""),
        ("Freno",freno,"",""),
        ("Puntos de muestreo",len(df_f),"registros","kpi-ok"),
        ("Planicidad inicial",f"{avg_f_val:.3f} mm",str(r_first['año']),""),
        ("Planicidad actual",f"{avg_l_val:.3f} mm",str(r_last['año']),
         "kpi-warn" if avg_l_val>threshold else "kpi-ok"),
        ("Variación acumulada",f"{wear:.3f} mm","primera → última medición",
         "kpi-warn" if abs(wear)>0.5 else ""),
        ("Tasa anual",f"{rate_yr:.3f} mm/a","promedio histórico",
         "kpi-warn" if abs(rate_yr)>0.2 else ""),
    ])

    tab1,tab2,tab3=st.tabs(["📏 Evolución temporal","📈 Regresión & Pronóstico","🌡️ Mapa de calor"])

    with tab1:
        st.markdown('<div class="section-title">Evolución de planicidad en los 4 puntos de medida</div>',
                    unsafe_allow_html=True)
        fig1=go.Figure()
        for col,lbl in FRENO_LABELS.items():
            if col not in df_f.columns: continue
            fig1.add_trace(go.Scatter(x=df_f["año"].astype(str),y=df_f[col],
                mode="lines+markers",name=lbl,
                line=dict(color=FRENO_COLORS[col],width=2),marker=dict(size=9),
                hovertemplate=f"{lbl}<br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
        fig1.add_hline(y=FRENO_REF,line_dash="dot",line_color="#546e8a",
                       annotation_text=f"Nominal {FRENO_REF} mm",annotation_font_color="#546e8a")
        fig1.update_layout(**BL,height=400,
            title=dict(text=f"📏 Planicidad Pista de Freno — {unit} · {freno}",font_size=14,x=0.01),
            xaxis_title="Año",yaxis_title="Planicidad (mm)")
        st.plotly_chart(fig1,use_container_width=True)
        disp=df_f[["año","freno"]+FRENO_COLS].copy()
        disp.columns=["Año","Freno"]+list(FRENO_LABELS.values())
        st.dataframe(disp.style.background_gradient(cmap="RdYlGn_r",
                     subset=list(FRENO_LABELS.values())),use_container_width=True,height=260)

    with tab2:
        st.markdown(f'<div class="section-title">Regresión — {unit} · {freno} · {punto_reg}</div>',
                    unsafe_allow_html=True)
        if col_reg in df_f.columns:
            df_s=df_f.sort_values("fecha")
            x,t0=days_from_origin(df_s["fecha"])
            y=df_s[col_reg].values.astype(float)
            valid=~np.isnan(y)
            if valid.sum()>=2:
                x_v=x[valid]; y_v=y[valid]
                all_fits=get_all_fits(x_v,y_v)
                sel_fits=[f for f in all_fits if f["name"] in reg_sel] if reg_sel else all_fits
                fig3=go.Figure()
                fig3.add_trace(go.Scatter(x=df_s["fecha"].dt.strftime("%Y-%m-%d"),y=y,
                    mode="markers",name="Medición real",
                    marker=dict(size=11,color="#ffd54f",symbol="circle",
                                line=dict(color="#ff8f00",width=1.5)),
                    hovertemplate="Año: %{x}<br>%{y:.3f} mm<extra></extra>"))
                forecasts=regression_block(fig3,x_v,t0,y_v,sel_fits,threshold,forecast_yr)
                fig3.update_layout(**BL,height=460,
                    title=dict(text=f"📈 Regresión — {unit} · {freno} · {punto_reg}",
                               font_size=14,x=0.01),xaxis_title="Fecha",yaxis_title="Planicidad (mm)")
                st.plotly_chart(fig3,use_container_width=True)
                show_forecasts(forecasts,threshold,forecast_yr)
            else:
                st.info("Se necesitan al menos 2 puntos válidos para la regresión.")

        st.markdown('<div class="section-title">Mejor modelo por punto de medida</div>',unsafe_allow_html=True)
        x_d,_=days_from_origin(df_f.sort_values("fecha")["fecha"])
        rows_tbl=[]
        for col,lbl in FRENO_LABELS.items():
            if col not in df_f.columns: continue
            yc=df_f.sort_values("fecha")[col].values.astype(float)
            valid=~np.isnan(yc)
            if valid.sum()<2: continue
            best=get_all_fits(x_d[valid],yc[valid])[0]
            rows_tbl.append({"Punto":lbl,"Mejor modelo":best["name"],
                              "R²":round(best["r2"],4),"RMSE":round(best["rmse"],4),
                              "Valor actual (mm)":round(float(df_f.sort_values("año")[col].iloc[-1]),3)})
        if rows_tbl:
            st.dataframe(pd.DataFrame(rows_tbl).style.map(color_r2,subset=["R²"]),
                         use_container_width=True,height=220)

    with tab3:
        st.markdown('<div class="section-title">Mapa de calor — todos los frenos × puntos (último año)</div>',
                    unsafe_allow_html=True)
        last_yr=df_unit["año"].max()
        df_last=df_unit[df_unit["año"]==last_yr]
        z_hm=[[row.get(c,np.nan) for c in FRENO_COLS] for _,row in df_last.iterrows()]
        fig_hm=go.Figure(go.Heatmap(z=z_hm,x=list(FRENO_LABELS.values()),
            y=df_last["freno"].tolist(),colorscale="RdYlGn_r",
            text=[[f"{v:.2f}" for v in row] for row in z_hm],
            texttemplate="%{text}",textfont=dict(size=10,color="#000"),
            hovertemplate="%{y} · %{x}<br>%{z:.3f} mm<extra></extra>",
            colorbar=dict(title=dict(text="mm",font=dict(color="#c8d0e0")),tickfont=dict(color="#c8d0e0"))))
        fig_hm.update_layout(**BL,height=400,
            title=dict(text=f"🌡️ Mapa de calor — {unit} · {last_yr}",font_size=14,x=0.01),
            xaxis_title="Punto de medida",yaxis_title="Freno")
        st.plotly_chart(fig_hm,use_container_width=True)

# ══════════════════════════════════════════════════════════════
# MÓDULO 4 — PLACA SELLO EJE
# ══════════════════════════════════════════════════════════════
SELLO_NOMINAL=25.5

def mod_sello(data):
    keys={k:v for k,v in data.items() if k.startswith("sello_")}
    if not keys:
        st.warning("No hay datos de Placa Sello Eje disponibles."); return

    unit_map={k.replace("sello_",""):k for k in keys}
    unit=st.sidebar.selectbox("🔧 Unidad",list(unit_map.keys()),key="sello_unit")
    df_unit=keys[unit_map[unit]]
    puntos=sorted(df_unit["punto"].unique())
    punto=st.sidebar.selectbox("📍 Punto angular",puntos,key="sello_punto")
    df_p=df_unit[df_unit["punto"]==punto].sort_values("año").reset_index(drop=True)
    df_p["fecha"]=pd.to_datetime(df_p["año"].astype(str)+"-01-01")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📊 Regresión")
    sensor=st.sidebar.radio("Sensor",["Sensor 1","Sensor 2"],key="sello_sensor")
    col_reg="sensor_1" if sensor=="Sensor 1" else "sensor_2"
    reg_opts=["Lineal","Polinómica g2","Polinómica g3","Exponencial","Potencial"]
    reg_sel=st.sidebar.multiselect("Modelos",reg_opts,
                                    default=["Lineal","Polinómica g2","Exponencial"],key="sello_reg")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚠️ Pronóstico")
    threshold=st.sidebar.number_input("Espesor límite crítico (mm)",20.0,30.0,SELLO_NOMINAL,0.01,
                                       key="sello_thr")
    forecast_yr=st.sidebar.slider("Horizonte (años)",1,10,4,key="sello_fc")

    avg_f=np.nanmean([df_p.iloc[0].get("sensor_1",np.nan),df_p.iloc[0].get("sensor_2",np.nan)])
    avg_l=np.nanmean([df_p.iloc[-1].get("sensor_1",np.nan),df_p.iloc[-1].get("sensor_2",np.nan)])
    wear=avg_f-avg_l
    n_yrs=df_p.iloc[-1]["año"]-df_p.iloc[0]["año"]
    rate_yr=wear/n_yrs if n_yrs>0 else 0

    kpi_row([
        ("Unidad",unit,"",""),
        ("Punto angular",punto,"",""),
        ("Puntos de muestreo",len(df_p),"registros","kpi-ok"),
        ("Espesor inicial prom",f"{avg_f:.3f} mm",str(df_p.iloc[0]['año']),""),
        ("Espesor actual prom",f"{avg_l:.3f} mm",str(df_p.iloc[-1]['año']),
         "kpi-warn" if avg_l<threshold else "kpi-ok"),
        ("Variación acumulada",f"{wear:.3f} mm","primera → última medición",
         "kpi-warn" if abs(wear)>0.1 else ""),
        ("Tasa anual",f"{rate_yr:.3f} mm/a","promedio histórico",
         "kpi-warn" if abs(rate_yr)>0.05 else ""),
    ])

    tab1,tab2,tab3=st.tabs(["📏 Evolución temporal","📈 Regresión & Pronóstico","🌡️ Mapa de calor"])

    with tab1:
        st.markdown('<div class="section-title">Evolución del espesor por año</div>',unsafe_allow_html=True)
        fig1=go.Figure()
        for col,lbl,clr in [("sensor_1","Sensor 1","#42a5f5"),("sensor_2","Sensor 2","#ffa726")]:
            if col not in df_p.columns: continue
            fig1.add_trace(go.Scatter(x=df_p["año"].astype(str),y=df_p[col],
                mode="lines+markers",name=lbl,
                line=dict(color=clr,width=2),marker=dict(size=9),
                hovertemplate=f"{lbl}<br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
        fig1.add_hline(y=SELLO_NOMINAL,line_dash="dot",line_color="#546e8a",
                       annotation_text=f"Nominal {SELLO_NOMINAL} mm",annotation_font_color="#546e8a")
        fig1.update_layout(**BL,height=400,
            title=dict(text=f"📏 Placa Sello Eje — {unit} · {punto}",font_size=14,x=0.01),
            xaxis_title="Año",yaxis_title="Espesor (mm)")
        st.plotly_chart(fig1,use_container_width=True)
        disp=df_p[["año","punto","sensor_1","sensor_2"]].copy()
        disp.columns=["Año","Punto","Sensor 1 (mm)","Sensor 2 (mm)"]
        st.dataframe(disp,use_container_width=True,height=200)

    with tab2:
        st.markdown(f'<div class="section-title">Regresión — {unit} · {punto} · {sensor}</div>',
                    unsafe_allow_html=True)
        if col_reg in df_p.columns:
            df_s=df_p.sort_values("fecha")
            x,t0=days_from_origin(df_s["fecha"])
            y=df_s[col_reg].values.astype(float)
            valid=~np.isnan(y)
            if valid.sum()>=2:
                all_fits=get_all_fits(x[valid],y[valid])
                sel_fits=[f for f in all_fits if f["name"] in reg_sel] if reg_sel else all_fits
                fig3=go.Figure()
                fig3.add_trace(go.Scatter(x=df_s["fecha"].dt.strftime("%Y-%m-%d"),y=y,
                    mode="markers",name="Medición real",
                    marker=dict(size=11,color="#ffd54f",symbol="circle",
                                line=dict(color="#ff8f00",width=1.5)),
                    hovertemplate="Año: %{x}<br>%{y:.3f} mm<extra></extra>"))
                forecasts=regression_block(fig3,x[valid],t0,y[valid],sel_fits,threshold,forecast_yr)
                fig3.update_layout(**BL,height=460,
                    title=dict(text=f"📈 Regresión — {unit} · {punto} · {sensor}",font_size=14,x=0.01),
                    xaxis_title="Fecha",yaxis_title="Espesor (mm)")
                st.plotly_chart(fig3,use_container_width=True)
                show_forecasts(forecasts,threshold,forecast_yr)

    with tab3:
        st.markdown('<div class="section-title">Mapa de calor — todos los puntos angulares (último año)</div>',
                    unsafe_allow_html=True)
        last_yr=df_unit["año"].max()
        df_last=df_unit[df_unit["año"]==last_yr].sort_values("punto")
        z_hm=[[row.get("sensor_1",np.nan),row.get("sensor_2",np.nan)] for _,row in df_last.iterrows()]
        fig_hm=go.Figure(go.Heatmap(z=z_hm,x=["Sensor 1","Sensor 2"],
            y=df_last["punto"].tolist(),colorscale="RdYlGn_r",
            text=[[f"{v:.3f}" for v in row] for row in z_hm],
            texttemplate="%{text}",textfont=dict(size=10,color="#000"),
            hovertemplate="%{y} · %{x}<br>%{z:.3f} mm<extra></extra>",
            colorbar=dict(title=dict(text="mm",font=dict(color="#c8d0e0")),tickfont=dict(color="#c8d0e0"))))
        fig_hm.update_layout(**BL,height=400,
            title=dict(text=f"🌡️ Mapa de calor — {unit} · {last_yr}",font_size=14,x=0.01),
            xaxis_title="Sensor",yaxis_title="Punto angular")
        st.plotly_chart(fig_hm,use_container_width=True)

# ══════════════════════════════════════════════════════════════
# MÓDULO 5 — COJINETE GUÍA
# ══════════════════════════════════════════════════════════════
def mod_cojinete(data):
    keys={k:v for k,v in data.items() if k.startswith("cojinete_")}
    if not keys:
        st.warning("No hay datos de Cojinete Guía disponibles."); return

    unit_map={k.replace("cojinete_",""):k for k in keys}
    unit=st.sidebar.selectbox("🔧 Unidad",list(unit_map.keys()),key="coj_unit")
    df_unit=keys[unit_map[unit]]
    puntos=sorted(df_unit["punto"].unique())
    punto=st.sidebar.selectbox("📍 Punto (I–VIII)",puntos,key="coj_punto")
    df_p=df_unit[df_unit["punto"]==punto].sort_values("año").reset_index(drop=True)
    df_p["fecha"]=pd.to_datetime(df_p["año"].astype(str)+"-01-01")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📊 Regresión")
    reg_opts=["Lineal","Polinómica g2","Polinómica g3","Exponencial","Potencial"]
    reg_sel=st.sidebar.multiselect("Modelos",reg_opts,
                                    default=["Lineal","Polinómica g2","Exponencial"],key="coj_reg")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚠️ Pronóstico")
    threshold=st.sidebar.number_input("Holgura límite crítica (mm)",0.0,2.0,0.5,0.01,key="coj_thr")
    forecast_yr=st.sidebar.slider("Horizonte (años)",1,10,4,key="coj_fc")

    avg_f=df_p.iloc[0].get("holgura",np.nan)
    avg_l=df_p.iloc[-1].get("holgura",np.nan)
    wear=float(avg_f-avg_l) if not np.isnan(avg_f) and not np.isnan(avg_l) else 0
    n_yrs=df_p.iloc[-1]["año"]-df_p.iloc[0]["año"]
    rate_yr=wear/n_yrs if n_yrs>0 else 0

    kpi_row([
        ("Unidad",unit,"",""),
        ("Punto",punto,"",""),
        ("Puntos de muestreo",len(df_p),"registros","kpi-ok"),
        ("Holgura inicial",f"{avg_f:.3f} mm",str(df_p.iloc[0]['año']),""),
        ("Holgura actual",f"{avg_l:.3f} mm",str(df_p.iloc[-1]['año']),
         "kpi-warn" if avg_l>threshold else "kpi-ok"),
        ("Variación acumulada",f"{wear:.3f} mm","primera → última medición",
         "kpi-warn" if abs(wear)>0.1 else ""),
        ("Tasa anual",f"{rate_yr:.3f} mm/a","promedio histórico",
         "kpi-warn" if abs(rate_yr)>0.05 else ""),
    ])

    tab1,tab2,tab3=st.tabs(["📏 Evolución temporal","📈 Regresión & Pronóstico","🌡️ Mapa de calor"])

    with tab1:
        st.markdown('<div class="section-title">Evolución de la holgura por año</div>',unsafe_allow_html=True)
        fig1=go.Figure()
        fig1.add_trace(go.Scatter(x=df_p["año"].astype(str),y=df_p["holgura"],
            mode="lines+markers",name=f"Punto {punto}",
            line=dict(color="#42a5f5",width=2),marker=dict(size=9),
            hovertemplate=f"Punto {punto}<br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
        if threshold>0:
            fig1.add_hline(y=threshold,line_dash="dot",line_color="#ef5350",
                           annotation_text=f"Límite {threshold} mm",annotation_font_color="#ef5350")
        fig1.update_layout(**BL,height=380,
            title=dict(text=f"📏 Holgura Cojinete Guía — {unit} · Punto {punto}",font_size=14,x=0.01),
            xaxis_title="Año",yaxis_title="Holgura (mm)")
        st.plotly_chart(fig1,use_container_width=True)

        # Todos los puntos juntos
        st.markdown('<div class="section-title">Todos los puntos (I–VIII)</div>',unsafe_allow_html=True)
        fig_all=go.Figure()
        for i,p in enumerate(puntos):
            sub=df_unit[df_unit["punto"]==p].sort_values("año")
            fig_all.add_trace(go.Scatter(x=sub["año"].astype(str),y=sub["holgura"],
                mode="lines+markers",name=f"Punto {p}",
                line=dict(color=PAL[i%len(PAL)],width=2),marker=dict(size=7),
                hovertemplate=f"Punto {p}<br>%{{x}}<br>%{{y:.3f}} mm<extra></extra>"))
        fig_all.update_layout(**BL,height=400,
            title=dict(text=f"📏 Todos los puntos — {unit}",font_size=14,x=0.01),
            xaxis_title="Año",yaxis_title="Holgura (mm)")
        st.plotly_chart(fig_all,use_container_width=True)
        disp=df_p[["año","punto","holgura"]].copy()
        disp.columns=["Año","Punto","Holgura (mm)"]
        st.dataframe(disp,use_container_width=True,height=180)

    with tab2:
        st.markdown(f'<div class="section-title">Regresión — {unit} · Punto {punto}</div>',
                    unsafe_allow_html=True)
        df_s=df_p.sort_values("fecha")
        x,t0=days_from_origin(df_s["fecha"])
        y=df_s["holgura"].values.astype(float)
        valid=~np.isnan(y)
        if valid.sum()>=2:
            all_fits=get_all_fits(x[valid],y[valid])
            sel_fits=[f for f in all_fits if f["name"] in reg_sel] if reg_sel else all_fits
            fig3=go.Figure()
            fig3.add_trace(go.Scatter(x=df_s["fecha"].dt.strftime("%Y-%m-%d"),y=y,
                mode="markers",name="Medición real",
                marker=dict(size=11,color="#ffd54f",symbol="circle",
                            line=dict(color="#ff8f00",width=1.5)),
                hovertemplate="Año: %{x}<br>%{y:.3f} mm<extra></extra>"))
            forecasts=regression_block(fig3,x[valid],t0,y[valid],sel_fits,threshold,forecast_yr)
            fig3.update_layout(**BL,height=460,
                title=dict(text=f"📈 Regresión — {unit} · Punto {punto}",font_size=14,x=0.01),
                xaxis_title="Fecha",yaxis_title="Holgura (mm)")
            st.plotly_chart(fig3,use_container_width=True)
            show_forecasts(forecasts,threshold,forecast_yr)
        else:
            st.info("Se necesitan al menos 2 puntos válidos para la regresión.")

    with tab3:
        st.markdown('<div class="section-title">Mapa de calor — todos los puntos × años</div>',
                    unsafe_allow_html=True)
        years_all=sorted(df_unit["año"].unique())
        z_hm=[]
        for p in puntos:
            row_vals=[]
            for yr in years_all:
                v=df_unit[(df_unit["punto"]==p)&(df_unit["año"]==yr)]["holgura"]
                row_vals.append(float(v.iloc[0]) if len(v)>0 else np.nan)
            z_hm.append(row_vals)
        fig_hm=go.Figure(go.Heatmap(z=z_hm,x=[str(y) for y in years_all],y=puntos,
            colorscale="RdYlGn_r",
            text=[[f"{v:.3f}" if not np.isnan(v) else "" for v in row] for row in z_hm],
            texttemplate="%{text}",textfont=dict(size=10,color="#000"),
            hovertemplate="Punto %{y} · %{x}<br>%{z:.3f} mm<extra></extra>",
            colorbar=dict(title=dict(text="mm",font=dict(color="#c8d0e0")),tickfont=dict(color="#c8d0e0"))))
        fig_hm.update_layout(**BL,height=380,
            title=dict(text=f"🌡️ Mapa de calor — {unit}",font_size=14,x=0.01),
            xaxis_title="Año",yaxis_title="Punto")
        st.plotly_chart(fig_hm,use_container_width=True)

# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
MODULOS={
    "⚙️ Rodete":       mod_rodete,
    "🔩 Directrices":  mod_directriz,
    "🛑 Pista de Freno": mod_freno,
    "🔧 Placa Sello Eje": mod_sello,
    "🎯 Cojinete Guía":  mod_cojinete,
}

def main():
    st.markdown("""
    <div class="ge-header">
      <div>
        <h1>⚡ GE Vernova — Monitoreo de Desgaste</h1>
        <span>Rodete · Directrices · Pista de Freno · Placa Sello Eje · Cojinete Guía</span>
      </div>
    </div>""", unsafe_allow_html=True)

    sidebar_data_panel()
    data=get_data()

    if not data:
        st.warning("⚠️ No hay datos. Sube el Excel en el panel lateral o agrega los CSV a /data.")
        st.stop()

    with st.sidebar:
        st.markdown("### 🗂️ Módulo de análisis")
        modulo=st.radio("",list(MODULOS.keys()),label_visibility="collapsed",key="modulo")
        st.markdown("---")
        st.markdown("### ⚙️ Panel de Control")
        st.markdown("---")

    MODULOS[modulo](data)

if __name__=="__main__":
    main()
